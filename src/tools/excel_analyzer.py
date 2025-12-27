"""Excel file analysis tools using openpyxl, formulas, and oletools."""

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula import Tokenizer

from src.models import (
    CellInfo,
    FormulaInfo,
    VBAModule,
    SheetInfo,
    PrintSettings,
    ExcelAnalysis,
)


# Paper size mapping (openpyxl paperSize values)
PAPER_SIZES = {
    1: "Letter",
    5: "Legal",
    8: "A3",
    9: "A4",
    11: "A5",
}


def analyze_excel_file(file_path: str) -> ExcelAnalysis:
    """
    Analyze an Excel file and extract all structural information.

    Args:
        file_path: Path to the Excel file (.xlsx, .xlsm, .xls)

    Returns:
        ExcelAnalysis model with complete file analysis
    """
    path = Path(file_path)
    file_type = path.suffix.lower().lstrip(".")

    # Load workbook (data_only=False to get formulas)
    wb = load_workbook(file_path, data_only=False)

    # Analyze each sheet
    sheets = []
    all_formulas = []
    all_input_cells = []
    all_output_cells = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = _analyze_sheet(ws)
        sheets.append(sheet_info)
        all_formulas.extend(sheet_info.formulas)
        all_input_cells.extend(sheet_info.input_cells)
        all_output_cells.extend(sheet_info.output_cells)

    # Extract VBA if present
    vba_modules = []
    has_vba = file_type == "xlsm"
    if has_vba:
        vba_modules = _extract_vba(file_path)

    # Get print settings from first sheet
    print_settings = _extract_print_settings(wb.active) if wb.active else None

    # Calculate complexity score
    complexity = _calculate_complexity(
        len(all_formulas),
        len(vba_modules),
        len(sheets)
    )

    wb.close()

    return ExcelAnalysis(
        filename=path.name,
        file_type=file_type,
        sheets=sheets,
        vba_modules=vba_modules,
        has_vba=has_vba,
        print_settings=print_settings,
        total_formulas=len(all_formulas),
        total_input_cells=len(all_input_cells),
        total_output_cells=len(all_output_cells),
        complexity_score=complexity,
    )


def _analyze_sheet(ws: Worksheet) -> SheetInfo:
    """Analyze a single worksheet."""
    # Get dimensions
    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    used_range = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"

    # Find all formulas
    formulas = []
    formula_cells = set()

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.data_type == "f" and cell.value:
                formula_str = str(cell.value)
                deps = _extract_cell_references(formula_str)
                result_type = _infer_formula_result_type(formula_str)

                formulas.append(FormulaInfo(
                    cell=cell.coordinate,
                    formula=formula_str,
                    dependencies=deps,
                    result_type=result_type,
                ))
                formula_cells.add(cell.coordinate)

    # Detect input cells (referenced by formulas but not formulas themselves)
    referenced_cells = set()
    for formula in formulas:
        referenced_cells.update(formula.dependencies)

    input_cells = list(referenced_cells - formula_cells)
    output_cells = [f.cell for f in formulas]

    # Check print area
    print_area = ws.print_area
    has_print_area = bool(print_area)

    return SheetInfo(
        name=ws.title,
        row_count=max_row,
        col_count=max_col,
        used_range=used_range,
        input_cells=sorted(input_cells),
        output_cells=sorted(output_cells),
        formulas=formulas,
        has_print_area=has_print_area,
        print_area=print_area if has_print_area else None,
    )


def _extract_cell_references(formula: str) -> list[str]:
    """Extract cell references from a formula using openpyxl tokenizer."""
    try:
        tokenizer = Tokenizer(formula)
        refs = []
        for token in tokenizer.items:
            if token.type == "OPERAND" and token.subtype == "RANGE":
                # Handle range references like A1:B10
                value = token.value
                # Remove sheet reference if present
                if "!" in value:
                    value = value.split("!")[-1]
                # Remove absolute reference markers
                value = value.replace("$", "")

                if ":" in value:
                    # Expand range to individual cells
                    refs.extend(_expand_range(value))
                else:
                    refs.append(value)
        return list(set(refs))
    except Exception:
        # Fallback: simple regex for cell references
        pattern = r"[A-Z]+[0-9]+"
        return list(set(re.findall(pattern, formula.upper())))


def _expand_range(range_str: str) -> list[str]:
    """Expand a range like A1:B3 to individual cells."""
    try:
        start, end = range_str.split(":")
        start_col = re.match(r"([A-Z]+)", start).group(1)
        start_row = int(re.search(r"(\d+)", start).group(1))
        end_col = re.match(r"([A-Z]+)", end).group(1)
        end_row = int(re.search(r"(\d+)", end).group(1))

        cells = []
        for col_idx in range(_col_to_num(start_col), _col_to_num(end_col) + 1):
            for row in range(start_row, end_row + 1):
                cells.append(f"{_num_to_col(col_idx)}{row}")
        return cells
    except Exception:
        return [range_str]


def _col_to_num(col: str) -> int:
    """Convert column letter to number (A=1, B=2, etc.)."""
    result = 0
    for char in col:
        result = result * 26 + (ord(char.upper()) - ord("A") + 1)
    return result


def _num_to_col(num: int) -> str:
    """Convert number to column letter (1=A, 2=B, etc.)."""
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _infer_formula_result_type(formula: str) -> str:
    """Infer the result type of a formula based on the function used."""
    formula_upper = formula.upper()

    # String functions
    string_funcs = ["CONCATENATE", "LEFT", "RIGHT", "MID", "TEXT", "UPPER", "LOWER"]
    if any(func in formula_upper for func in string_funcs):
        return "string"

    # Boolean functions
    bool_funcs = ["AND", "OR", "NOT", "ISBLANK", "ISERROR", "TRUE", "FALSE"]
    if any(func in formula_upper for func in bool_funcs):
        return "boolean"

    # Date functions
    date_funcs = ["DATE", "TODAY", "NOW", "YEAR", "MONTH", "DAY"]
    if any(func in formula_upper for func in date_funcs):
        return "date"

    # Default to number (most common)
    return "number"


def _extract_vba(file_path: str) -> list[VBAModule]:
    """Extract VBA modules from an Excel file using oletools."""
    try:
        from oletools.olevba import VBA_Parser

        vba_parser = VBA_Parser(file_path)
        modules = []

        if vba_parser.detect_vba_macros():
            for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                # Determine module type
                if "Class" in vba_filename:
                    module_type = "Class"
                elif "Sheet" in vba_filename or "ThisWorkbook" in vba_filename:
                    module_type = "Sheet"
                elif "UserForm" in vba_filename:
                    module_type = "Form"
                else:
                    module_type = "Module"

                # Extract procedure names
                procedures = _extract_vba_procedures(vba_code)

                modules.append(VBAModule(
                    name=vba_filename,
                    module_type=module_type,
                    code=vba_code,
                    procedures=procedures,
                ))

        vba_parser.close()
        return modules

    except ImportError:
        return []
    except Exception:
        return []


def _extract_vba_procedures(code: str) -> list[str]:
    """Extract procedure and function names from VBA code."""
    procedures = []

    # Match Sub and Function declarations
    pattern = r"(?:Public\s+|Private\s+)?(?:Sub|Function)\s+(\w+)"
    matches = re.findall(pattern, code, re.IGNORECASE)
    procedures.extend(matches)

    return procedures


def _extract_print_settings(ws: Worksheet) -> PrintSettings:
    """Extract print settings from a worksheet."""
    page_setup = ws.page_setup
    margins = ws.page_margins

    # Get orientation
    orientation = "landscape" if page_setup.orientation == "landscape" else "portrait"

    # Get paper size
    paper_size = PAPER_SIZES.get(page_setup.paperSize, "A4")

    # Get margins (convert to inches if needed)
    margin_dict = {
        "top": float(margins.top) if margins.top else 0.75,
        "bottom": float(margins.bottom) if margins.bottom else 0.75,
        "left": float(margins.left) if margins.left else 0.7,
        "right": float(margins.right) if margins.right else 0.7,
    }

    return PrintSettings(
        orientation=orientation,
        paper_size=paper_size,
        margins=margin_dict,
        header=ws.oddHeader.center.text if ws.oddHeader and ws.oddHeader.center else None,
        footer=ws.oddFooter.center.text if ws.oddFooter and ws.oddFooter.center else None,
        fit_to_page=page_setup.fitToPage or False,
        scale=page_setup.scale or 100,
    )


def _calculate_complexity(formula_count: int, vba_count: int, sheet_count: int) -> str:
    """Calculate overall complexity score."""
    score = 0

    # Formula complexity
    if formula_count > 50:
        score += 3
    elif formula_count > 20:
        score += 2
    elif formula_count > 0:
        score += 1

    # VBA complexity (weighs heavily)
    if vba_count > 0:
        score += 3

    # Sheet complexity
    if sheet_count > 5:
        score += 2
    elif sheet_count > 2:
        score += 1

    if score >= 5:
        return "high"
    elif score >= 2:
        return "medium"
    else:
        return "low"


def get_cell_data(file_path: str, sheet_name: str = None) -> dict[str, CellInfo]:
    """
    Get detailed cell data from a worksheet.

    Args:
        file_path: Path to the Excel file
        sheet_name: Optional sheet name (defaults to active sheet)

    Returns:
        Dictionary mapping cell addresses to CellInfo
    """
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None or cell.data_type == "f":
                data_type = _get_cell_data_type(cell)
                cells[cell.coordinate] = CellInfo(
                    address=cell.coordinate,
                    value=cell.value,
                    formula=str(cell.value) if cell.data_type == "f" else None,
                    data_type=data_type,
                    format=cell.number_format,
                )

    wb.close()
    return cells


def _get_cell_data_type(cell: Cell) -> str:
    """Get the data type of a cell."""
    if cell.data_type == "f":
        return "formula"
    elif cell.data_type == "n":
        return "number"
    elif cell.data_type == "s":
        return "string"
    elif cell.data_type == "b":
        return "boolean"
    elif cell.data_type == "d":
        return "date"
    elif cell.value is None:
        return "empty"
    else:
        return "string"
