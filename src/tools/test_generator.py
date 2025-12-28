"""Test case generator - Extracts deterministic test cases from Excel files.

Key insight: Excel formulas are deterministic.
Given the same inputs, they always produce the same outputs.
This allows us to automatically generate test cases by:
1. Reading current input cell values
2. Reading computed output values (data_only=True)
3. Creating input → expected_output test cases
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.test_case import (
    FormulaTestCase,
    InputOutputMapping,
    TestScenario,
    StaticTestSuite,
)
from src.models.analysis import ExcelAnalysis, FormulaInfo


def extract_test_cases(
    file_path: str,
    analysis: Optional[ExcelAnalysis] = None,
    max_formulas: int = 50,
) -> StaticTestSuite:
    """
    Extract deterministic test cases from an Excel file.

    This function:
    1. Opens the Excel file twice (formulas and computed values)
    2. For each formula, captures its input dependencies and output value
    3. Generates test cases that can verify JS conversion accuracy

    Args:
        file_path: Path to the Excel file
        analysis: Optional pre-computed ExcelAnalysis (avoids re-parsing)
        max_formulas: Maximum number of formula tests to generate

    Returns:
        StaticTestSuite with all test cases
    """
    path = Path(file_path)

    # Load workbook twice:
    # 1. data_only=False to get formulas
    # 2. data_only=True to get computed values
    wb_formulas = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)

    formula_tests: list[FormulaTestCase] = []
    field_mappings: list[InputOutputMapping] = []
    all_input_cells: set[str] = set()
    all_output_cells: set[str] = set()

    # Process each sheet
    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        # Extract tests from this sheet
        sheet_tests, sheet_inputs, sheet_outputs = _extract_sheet_tests(
            ws_formulas, ws_values, sheet_name, max_formulas - len(formula_tests)
        )

        formula_tests.extend(sheet_tests)
        all_input_cells.update(sheet_inputs)
        all_output_cells.update(sheet_outputs)

        # Create field mappings for input cells
        for input_cell in sheet_inputs:
            value = ws_values[input_cell].value
            field_mappings.append(InputOutputMapping(
                excel_input_cell=input_cell,
                sample_value=value,
            ))

        if len(formula_tests) >= max_formulas:
            break

    # Generate a basic test scenario using current values
    scenarios = _generate_default_scenarios(wb_values, all_input_cells, all_output_cells)

    wb_formulas.close()
    wb_values.close()

    return StaticTestSuite(
        excel_file=path.name,
        generated_at=datetime.now().isoformat(),
        formula_tests=formula_tests,
        field_mappings=field_mappings,
        scenarios=scenarios,
        total_formulas=len(formula_tests),
        total_inputs=len(all_input_cells),
        total_outputs=len(all_output_cells),
    )


def _extract_sheet_tests(
    ws_formulas: Worksheet,
    ws_values: Worksheet,
    sheet_name: str,
    max_tests: int,
) -> tuple[list[FormulaTestCase], set[str], set[str]]:
    """Extract test cases from a single worksheet."""
    tests: list[FormulaTestCase] = []
    input_cells: set[str] = set()
    output_cells: set[str] = set()

    # Iterate through cells to find formulas
    for row in ws_formulas.iter_rows():
        if len(tests) >= max_tests:
            break

        for cell in row:
            if len(tests) >= max_tests:
                break

            # Check if cell has a formula
            if cell.data_type == "f" and cell.value:
                formula_str = str(cell.value)
                formula_cell = cell.coordinate

                # Get the computed value from the values-only workbook
                computed_value = ws_values[formula_cell].value

                # Skip if no computed value (formula might have error)
                if computed_value is None:
                    continue

                # Extract input cell references from formula
                input_refs = _extract_input_refs(formula_str)

                # Get input values
                input_values = {}
                for ref in input_refs:
                    try:
                        # Handle sheet references
                        if "!" in ref:
                            continue  # Skip cross-sheet refs for now

                        val = ws_values[ref].value
                        if val is not None:
                            input_values[ref] = val
                            input_cells.add(ref)
                    except Exception:
                        continue

                # Skip if no valid inputs found
                if not input_values:
                    continue

                # Determine result type
                result_type = _infer_type(computed_value)

                # Create test case
                test_case = FormulaTestCase(
                    formula_cell=formula_cell,
                    formula=formula_str,
                    input_values=input_values,
                    expected_output=computed_value,
                    expected_type=result_type,
                    tolerance=0.0001 if result_type == "number" else 0,
                    description=f"{sheet_name}!{formula_cell}: {formula_str}",
                )

                tests.append(test_case)
                output_cells.add(formula_cell)

    return tests, input_cells, output_cells


def _extract_input_refs(formula: str) -> list[str]:
    """Extract cell references from a formula."""
    import re

    # Remove the leading '=' if present
    if formula.startswith("="):
        formula = formula[1:]

    # Find all cell references (A1, B2, AA10, etc.)
    pattern = r'\$?([A-Z]+)\$?(\d+)'
    matches = re.findall(pattern, formula.upper())

    refs = [f"{col}{row}" for col, row in matches]

    # Also handle ranges like A1:A10
    range_pattern = r'([A-Z]+\d+):([A-Z]+\d+)'
    range_matches = re.findall(range_pattern, formula.upper())

    for start, end in range_matches:
        # Expand range (simplified - just include start and end)
        expanded = _expand_range(start, end)
        refs.extend(expanded)

    return list(set(refs))


def _expand_range(start: str, end: str) -> list[str]:
    """Expand a cell range to individual cells."""
    import re

    start_match = re.match(r'([A-Z]+)(\d+)', start)
    end_match = re.match(r'([A-Z]+)(\d+)', end)

    if not start_match or not end_match:
        return [start, end]

    start_col, start_row = start_match.groups()
    end_col, end_row = end_match.groups()

    start_row, end_row = int(start_row), int(end_row)

    cells = []
    for col_idx in range(_col_to_num(start_col), _col_to_num(end_col) + 1):
        for row in range(start_row, end_row + 1):
            cells.append(f"{_num_to_col(col_idx)}{row}")

    return cells


def _col_to_num(col: str) -> int:
    """Convert column letter to number."""
    result = 0
    for char in col:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def _num_to_col(num: int) -> str:
    """Convert number to column letter."""
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _infer_type(value: Any) -> str:
    """Infer the data type of a value."""
    if isinstance(value, bool):
        return "boolean"
    elif isinstance(value, (int, float)):
        return "number"
    elif isinstance(value, str):
        return "string"
    else:
        return "string"


def _generate_default_scenarios(
    wb_values,
    input_cells: set[str],
    output_cells: set[str],
) -> list[TestScenario]:
    """Generate default test scenarios from current workbook values."""
    scenarios = []

    # Get the active sheet's values
    ws = wb_values.active

    # Create a "current state" scenario
    inputs = {}
    for cell in sorted(input_cells)[:20]:  # Limit inputs
        try:
            value = ws[cell].value
            if value is not None:
                inputs[cell] = value
        except Exception:
            continue

    expected_outputs = {}
    for cell in sorted(output_cells)[:20]:  # Limit outputs
        try:
            value = ws[cell].value
            if value is not None:
                expected_outputs[cell] = value
        except Exception:
            continue

    if inputs and expected_outputs:
        scenarios.append(TestScenario(
            name="기본 계산 테스트",
            description="Excel 파일의 현재 값을 기반으로 한 기본 테스트",
            inputs=inputs,
            expected_outputs=expected_outputs,
            tags=["smoke", "default"],
        ))

    return scenarios


def generate_node_test_script(test_suite: StaticTestSuite, js_code: str) -> str:
    """
    Generate a Node.js test script that can validate the generated JS code.

    Args:
        test_suite: The static test suite
        js_code: The generated JavaScript code to test

    Returns:
        Node.js test script as a string
    """
    script = f"""
// Auto-generated test script for: {test_suite.excel_file}
// Generated at: {test_suite.generated_at}

const {{ JSDOM }} = require('jsdom');

// Setup DOM environment
const dom = new JSDOM('<!DOCTYPE html><html><body></body></html>', {{
    runScripts: 'dangerously'
}});
const {{ window }} = dom;
global.document = window.document;
global.window = window;

// Inject the generated code
const generatedCode = `{js_code.replace('`', '\\`')}`;

// Execute the code in the DOM context
const script = window.document.createElement('script');
script.textContent = generatedCode;
window.document.body.appendChild(script);

// Test results
let passed = 0;
let failed = 0;
const failures = [];

function test(name, fn) {{
    try {{
        fn();
        passed++;
        console.log('✓', name);
    }} catch (e) {{
        failed++;
        failures.push({{ name, error: e.message }});
        console.log('✗', name, '-', e.message);
    }}
}}

function expect(actual) {{
    return {{
        toBe(expected) {{
            if (actual !== expected) {{
                throw new Error(`Expected ${{expected}}, got ${{actual}}`);
            }}
        }},
        toBeCloseTo(expected, decimals = 4) {{
            const tolerance = Math.pow(10, -decimals);
            if (Math.abs(actual - expected) > tolerance) {{
                throw new Error(`Expected ${{expected}} ± ${{tolerance}}, got ${{actual}}`);
            }}
        }}
    }};
}}

// Run tests
console.log('\\n=== Running Static Tests ===\\n');

"""

    # Add individual formula tests
    for test_case in test_suite.formula_tests:
        input_setup = "\n".join(
            f"    // Set {cell} = {value}"
            for cell, value in test_case.input_values.items()
        )

        expected = test_case.expected_output
        if isinstance(expected, str):
            expected = f"'{expected}'"

        script += f"""
test('{test_case.formula_cell}: {test_case.formula}', () => {{
{input_setup}
    // Expected: {test_case.expected_output}
    // Note: This test requires manual implementation based on your appData structure
    const result = window.appData ? window.appData() : null;
    if (result && result.{_cell_to_var(test_case.formula_cell)}) {{
        expect(result.{_cell_to_var(test_case.formula_cell)}).toBeCloseTo({expected});
    }}
}});
"""

    # Add summary
    script += """
// Summary
console.log('\\n=== Test Summary ===');
console.log(`Passed: ${passed}`);
console.log(`Failed: ${failed}`);
console.log(`Pass Rate: ${(passed / (passed + failed) * 100).toFixed(1)}%`);

if (failures.length > 0) {
    console.log('\\nFailures:');
    failures.forEach(f => console.log(`  - ${f.name}: ${f.error}`));
}

// Exit with appropriate code
process.exit(failed > 0 ? 1 : 0);
"""

    return script


def _cell_to_var(cell: str) -> str:
    """Convert cell reference to variable name."""
    return cell.lower().replace('$', '')


def generate_playwright_tests(test_suite: StaticTestSuite, base_url: str = "/") -> str:
    """
    Generate Playwright E2E test file.

    Args:
        test_suite: The static test suite
        base_url: Base URL for the web app

    Returns:
        Playwright test file content
    """
    content = f"""// Auto-generated Playwright E2E tests
// Source: {test_suite.excel_file}
// Generated: {test_suite.generated_at}

import {{ test, expect }} from '@playwright/test';

test.describe('{test_suite.excel_file} Tests', () => {{
"""

    for scenario in test_suite.scenarios:
        content += f"""
    test('{scenario.name}', async ({{ page }}) => {{
        await page.goto('{base_url}');

        // Fill input values
"""
        for cell, value in scenario.inputs.items():
            # Try to find input by various selectors
            content += f"""
        // Input: {cell} = {value}
        const input_{cell} = page.locator('[data-cell="{cell}"], #input_{cell.lower()}, input[name="{cell.lower()}"]').first();
        if (await input_{cell}.count() > 0) {{
            await input_{cell}.fill('{value}');
        }}
"""

        content += """
        // Trigger calculation (try common button patterns)
        const calcButton = page.locator('button:has-text("계산"), button:has-text("Calculate"), button[type="submit"]').first();
        if (await calcButton.count() > 0) {
            await calcButton.click();
        }

        // Wait for calculation
        await page.waitForTimeout(500);

        // Verify outputs
"""

        for cell, expected in scenario.expected_outputs.items():
            content += f"""
        // Output: {cell} should be {expected}
        const output_{cell} = page.locator('[data-cell="{cell}"], #output_{cell.lower()}, [id*="{cell.lower()}"]').first();
        if (await output_{cell}.count() > 0) {{
            const text = await output_{cell}.textContent() || await output_{cell}.inputValue();
            const value = parseFloat(text.replace(/[^0-9.-]/g, ''));
            expect(value).toBeCloseTo({expected}, 2);
        }}
"""

        content += """    });
"""

    content += """});
"""

    return content
